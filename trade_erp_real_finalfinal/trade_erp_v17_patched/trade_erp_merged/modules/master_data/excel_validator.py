# -*- coding: utf-8 -*-
"""
Excel 템플릿 검증 모듈
- PAGE2의 SUMIFS 수식 검증
- Excel 엔진을 통한 수식 재계산 (Windows 전용)
"""
import logging
import platform
from pathlib import Path
from typing import Dict, List, Any
import openpyxl

logger = logging.getLogger(__name__)


def verify_page2_formulas(filepath: str, verbose: bool = False) -> Dict[str, Any]:
    """
    PAGE2의 SUMIFS 수식 검증

    Args:
        filepath: Excel 파일 경로
        verbose: 상세 로그 출력 여부

    Returns:
        {
            'valid': bool,
            'formula_count': int,
            'formulas': List[Dict],
            'errors': List[str]
        }
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=False)

        if 'PAGE2_VIEW' not in wb.sheetnames:
            logger.warning("[VALIDATOR] PAGE2_VIEW 시트가 없습니다")
            return {
                'valid': False,
                'formula_count': 0,
                'formulas': [],
                'errors': ['PAGE2_VIEW 시트가 없습니다']
            }

        ws = wb['PAGE2_VIEW']

        formulas = []
        errors = []

        # 수식이 있는 셀 찾기 (일반적으로 집계 영역에 있음)
        for row in ws.iter_rows(min_row=2, max_row=100):  # 최대 100행까지 검사
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = {
                        'cell': cell.coordinate,
                        'formula': cell.value,
                        'valid': True
                    }

                    # SUMIFS가 PAGE1_DATA를 참조하는지 확인
                    if 'SUMIFS' in cell.value.upper():
                        if 'PAGE1_DATA' not in cell.value:
                            formula['valid'] = False
                            errors.append(f"{cell.coordinate}: PAGE1_DATA 참조 누락")
                    else:
                        # SUMIFS 외의 수식도 기록
                        if verbose:
                            logger.debug(f"[VALIDATOR] 비-SUMIFS 수식: {cell.coordinate}")

                    formulas.append(formula)

        wb.close()

        # 결과 요약
        valid = len(errors) == 0
        formula_count = len(formulas)

        if verbose:
            logger.info(f"[VALIDATOR] 수식 검증 완료: {formula_count}개 수식, {len(errors)}개 오류")
            for f in formulas[:5]:  # 처음 5개만 출력
                logger.info(f"  {f['cell']}: {f['formula'][:50]}...")

        return {
            'valid': valid,
            'formula_count': formula_count,
            'formulas': formulas,
            'errors': errors
        }

    except Exception as e:
        logger.error(f"[VALIDATOR] 검증 실패: {e}")
        return {
            'valid': False,
            'formula_count': 0,
            'formulas': [],
            'errors': [str(e)]
        }


def recalculate_excel(filepath: str) -> bool:
    """
    Excel 엔진을 사용하여 수식 재계산 (Windows 전용)

    Args:
        filepath: Excel 파일 경로

    Returns:
        성공 여부

    Note:
        - openpyxl은 수식을 실행하지 않음
        - Windows에서만 win32com으로 Excel 엔진 호출 가능
        - 기타 OS에서는 사용자가 직접 Excel을 열어야 함
    """
    if platform.system() != "Windows":
        logger.warning("[VALIDATOR] Excel 재계산은 Windows에서만 지원됩니다")
        return False

    try:
        import win32com.client
        import os

        logger.info(f"[VALIDATOR] Excel 재계산 시작: {filepath}")

        # Excel 애플리케이션 시작
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # 백그라운드 실행
        excel.DisplayAlerts = False  # 경고 비활성화

        # 파일 열기
        abs_path = os.path.abspath(filepath)
        wb = excel.Workbooks.Open(abs_path)

        # 수식 재계산
        excel.Calculate()
        wb.RefreshAll()  # 모든 데이터 새로고침

        # 저장 및 닫기
        wb.Save()
        wb.Close()
        excel.Quit()

        logger.info("[VALIDATOR] Excel 재계산 완료")
        return True

    except ImportError:
        logger.error("[VALIDATOR] win32com 모듈이 설치되지 않았습니다")
        logger.info("[VALIDATOR] 설치: pip install pywin32")
        return False

    except Exception as e:
        logger.error(f"[VALIDATOR] Excel 재계산 실패: {e}")

        # Excel 프로세스 정리
        try:
            if 'excel' in locals():
                excel.Quit()
        except:
            pass

        return False


def get_page2_summary(filepath: str) -> Dict[str, Any]:
    """
    PAGE2의 집계 데이터 요약 조회

    Args:
        filepath: Excel 파일 경로

    Returns:
        {
            'has_page2': bool,
            'row_count': int,
            'col_count': int,
            'sample_data': List[Dict]
        }
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)

        if 'PAGE2_VIEW' not in wb.sheetnames:
            return {
                'has_page2': False,
                'row_count': 0,
                'col_count': 0,
                'sample_data': []
            }

        ws = wb['PAGE2_VIEW']

        # 행/열 개수
        row_count = ws.max_row
        col_count = ws.max_column

        # 샘플 데이터 (처음 5행)
        sample_data = []
        for row_idx in range(1, min(6, row_count + 1)):
            row_data = {}
            for col_idx in range(1, col_count + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                row_data[f'col_{col_idx}'] = cell.value
            sample_data.append(row_data)

        wb.close()

        return {
            'has_page2': True,
            'row_count': row_count,
            'col_count': col_count,
            'sample_data': sample_data
        }

    except Exception as e:
        logger.error(f"[VALIDATOR] PAGE2 요약 실패: {e}")
        return {
            'has_page2': False,
            'row_count': 0,
            'col_count': 0,
            'sample_data': []
        }


if __name__ == "__main__":
    # 테스트 코드
    import sys

    if len(sys.argv) < 2:
        print("Usage: python excel_validator.py <excel_file_path>")
        sys.exit(1)

    filepath = sys.argv[1]

    print("=" * 60)
    print("Excel Validator Test")
    print("=" * 60)

    # 1. 수식 검증
    print("\n1. PAGE2 Formula Validation:")
    result = verify_page2_formulas(filepath, verbose=True)
    print(f"   Valid: {result['valid']}")
    print(f"   Formula Count: {result['formula_count']}")
    if result['errors']:
        print(f"   Errors: {result['errors']}")

    # 2. PAGE2 요약
    print("\n2. PAGE2 Summary:")
    summary = get_page2_summary(filepath)
    print(f"   Has PAGE2: {summary['has_page2']}")
    print(f"   Rows: {summary['row_count']}, Columns: {summary['col_count']}")

    # 3. 재계산 (Windows만)
    print("\n3. Excel Recalculation:")
    if platform.system() == "Windows":
        success = recalculate_excel(filepath)
        print(f"   Success: {success}")
    else:
        print(f"   Not supported on {platform.system()}")
