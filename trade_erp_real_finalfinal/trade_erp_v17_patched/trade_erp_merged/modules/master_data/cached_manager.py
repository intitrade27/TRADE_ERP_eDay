# -*- coding: utf-8 -*-
"""
캐시 기반 마스터 데이터 관리자
- 메모리 캐시(DataFrame) 기반 고속 CRUD
- ChangeTracker로 변경사항 추적
- Excel 파일과 양방향 동기화
- 타임스탬프 기반 충돌 해결
"""
import logging
import threading
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List, Optional
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

from modules.master_data.column_mapper import ColumnMapper
from modules.master_data.change_tracker import ChangeTracker, ChangeType

logger = logging.getLogger(__name__)


class CachedMasterDataManager:
    """
    메모리 캐시 기반 마스터 데이터 관리자

    아키텍처:
    - In-Memory Cache: pandas DataFrame (39 columns)
    - Change Tracker: 모든 CRUD 작업 추적
    - Column Mapper: 63개 → 39개 컬럼 변환
    - Excel Sync: PAGE1_DATA (rows 54-554)에 저장

    성능:
    - 읽기: <1ms (메모리)
    - 쓰기: <1ms (캐시) + 배치 동기화
    """

    # Excel 템플릿 설정
    EXCEL_SHEET_NAME = "PAGE1_DATA"
    EXCEL_HEADER_ROW = 53  # 헤더가 있는 행 (1-based)
    EXCEL_DATA_START_ROW = 54  # 데이터 시작 행 (1-based)
    EXCEL_DATA_END_ROW = 554  # 데이터 종료 행 (1-based, 500 lines)

    def __init__(
        self,
        excel_filepath: str,
        column_mapper: ColumnMapper,
        auto_load: bool = True
    ):
        """
        Args:
            excel_filepath: Excel 템플릿 파일 경로
            column_mapper: ColumnMapper 인스턴스
            auto_load: 초기화 시 Excel에서 자동 로드 여부
        """
        self.excel_filepath = Path(excel_filepath)
        self.mapper = column_mapper
        self.tracker = ChangeTracker()
        self.lock = threading.RLock()  # 재진입 가능 Lock

        # 캐시 초기화
        self.cache: Optional[pd.DataFrame] = None
        self._column_names: List[str] = []

        # Excel 파일 확인
        if not self.excel_filepath.exists():
            raise FileNotFoundError(f"Excel 템플릿 파일이 없습니다: {self.excel_filepath}")

        # 초기 로드
        if auto_load:
            self.load_from_excel()

        logger.info(f"[CACHED_MGR] 초기화 완료: {len(self.cache) if self.cache is not None else 0}행")

    def load_from_excel(self):
        """
        Excel 파일에서 캐시로 데이터 로드

        - PAGE1_DATA 시트의 rows 54-554 읽기
        - 39개 컬럼 구조로 DataFrame 생성
        - 빈 행은 제외
        """
        with self.lock:
            try:
                logger.info(f"[CACHED_MGR] Excel 로드 시작: {self.excel_filepath}")

                wb = openpyxl.load_workbook(self.excel_filepath, data_only=True)
                ws = wb[self.EXCEL_SHEET_NAME]

                # 헤더 읽기 (row 53)
                self._column_names = [
                    cell.value for cell in ws[self.EXCEL_HEADER_ROW]
                    if cell.value
                ]

                logger.info(f"[CACHED_MGR] 컬럼 수: {len(self._column_names)}")

                # 데이터 읽기 (rows 54-554)
                data_rows = []
                for row_idx in range(self.EXCEL_DATA_START_ROW, self.EXCEL_DATA_END_ROW + 1):
                    row_data = {}
                    is_empty = True

                    for col_idx, col_name in enumerate(self._column_names, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        value = cell.value

                        # 빈 값 체크
                        if value is not None and value != '':
                            is_empty = False

                        row_data[col_name] = value

                    # 빈 행이 아니면 추가
                    if not is_empty:
                        data_rows.append(row_data)

                wb.close()

                # DataFrame 생성
                if data_rows:
                    self.cache = pd.DataFrame(data_rows, columns=self._column_names)
                else:
                    self.cache = pd.DataFrame(columns=self._column_names)

                logger.info(f"[CACHED_MGR] 로드 완료: {len(self.cache)}행")

            except Exception as e:
                logger.error(f"[CACHED_MGR] Excel 로드 실패: {e}")
                raise

    def create_row(self, data: Dict[str, Any]) -> str:
        """
        캐시에 새 행 추가

        Args:
            data: 기존 63개 컬럼 데이터 (trade_id 포함)

        Returns:
            trade_id
        """
        with self.lock:
            try:
                trade_id = data.get('trade_id')
                if not trade_id:
                    raise ValueError("trade_id가 필요합니다")

                # 1. 컬럼 매핑 (63 → 39)
                mapped_data = self.mapper.apply_mapping(data)

                # 2. 타임스탬프 추가
                now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                # 컬럼명 정규화 (newline 제거)
                created_at_col = self._find_column('생성일시', 'created_at')
                updated_at_col = self._find_column('수정일시', 'updated_at')

                if created_at_col:
                    mapped_data[created_at_col] = now
                if updated_at_col:
                    mapped_data[updated_at_col] = now

                # 3. 캐시에 추가
                new_row = pd.DataFrame([mapped_data])
                self.cache = pd.concat([self.cache, new_row], ignore_index=True)

                # 4. 변경 추적
                self.tracker.track_create(trade_id, mapped_data)

                logger.info(f"[CACHED_MGR] 생성: {trade_id} (총 {len(self.cache)}행)")
                return trade_id

            except Exception as e:
                logger.error(f"[CACHED_MGR] 생성 실패: {e}")
                raise

    def update_row(self, trade_id: str, data: Dict[str, Any]) -> bool:
        """
        캐시에서 행 수정

        Args:
            trade_id: 거래 ID
            data: 수정할 필드 (63개 컬럼 구조)

        Returns:
            성공 여부
        """
        with self.lock:
            try:
                # 1. trade_id로 행 찾기
                trade_id_col = self._find_column('거래ID', 'trade_id')
                if not trade_id_col:
                    raise ValueError("거래ID 컬럼을 찾을 수 없습니다")

                mask = self.cache[trade_id_col] == trade_id
                if not mask.any():
                    logger.warning(f"[CACHED_MGR] 거래 없음: {trade_id}")
                    return False

                # 2. 컬럼 매핑
                mapped_data = self.mapper.apply_mapping(data)

                # 3. 타임스탬프 업데이트
                updated_at_col = self._find_column('수정일시', 'updated_at')
                if updated_at_col:
                    mapped_data[updated_at_col] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                # 4. 캐시 업데이트
                row_idx = self.cache[mask].index[0]
                for col, value in mapped_data.items():
                    if col in self.cache.columns:
                        self.cache.at[row_idx, col] = value

                # 5. 변경 추적
                self.tracker.track_update(trade_id, mapped_data, row_index=int(row_idx))

                logger.info(f"[CACHED_MGR] 수정: {trade_id}")
                return True

            except Exception as e:
                logger.error(f"[CACHED_MGR] 수정 실패: {e}")
                raise

    def delete_row(self, trade_id: str) -> bool:
        """
        캐시에서 행 삭제

        Args:
            trade_id: 거래 ID

        Returns:
            성공 여부
        """
        with self.lock:
            try:
                # 1. trade_id로 행 찾기
                trade_id_col = self._find_column('거래ID', 'trade_id')
                if not trade_id_col:
                    raise ValueError("거래ID 컬럼을 찾을 수 없습니다")

                mask = self.cache[trade_id_col] == trade_id
                if not mask.any():
                    logger.warning(f"[CACHED_MGR] 거래 없음: {trade_id}")
                    return False

                # 2. 행 인덱스 저장
                row_idx = self.cache[mask].index[0]

                # 3. 캐시에서 삭제
                self.cache = self.cache[~mask].reset_index(drop=True)

                # 4. 변경 추적
                self.tracker.track_delete(trade_id, row_index=int(row_idx))

                logger.info(f"[CACHED_MGR] 삭제: {trade_id} (남은 행: {len(self.cache)})")
                return True

            except Exception as e:
                logger.error(f"[CACHED_MGR] 삭제 실패: {e}")
                raise

    def read_all_rows(self) -> pd.DataFrame:
        """
        캐시에서 모든 행 읽기 (즉시 반환)

        Returns:
            DataFrame 복사본
        """
        with self.lock:
            return self.cache.copy() if self.cache is not None else pd.DataFrame()

    def read_row(self, trade_id: str) -> Optional[Dict[str, Any]]:
        """
        특정 거래 조회

        Args:
            trade_id: 거래 ID

        Returns:
            거래 데이터 딕셔너리 또는 None
        """
        with self.lock:
            trade_id_col = self._find_column('거래ID', 'trade_id')
            if not trade_id_col:
                return None

            mask = self.cache[trade_id_col] == trade_id
            if not mask.any():
                return None

            return self.cache[mask].iloc[0].to_dict()

    def sync_to_excel(self, force: bool = False):
        """
        변경사항을 Excel 파일에 반영

        Args:
            force: True면 전체 덮어쓰기, False면 변경사항만 반영
        """
        with self.lock:
            try:
                if force:
                    self._sync_full()
                else:
                    self._sync_incremental()

                # 변경사항 정리
                self.tracker.mark_all_synced()
                self.tracker.clear_synced()

                logger.info(f"[CACHED_MGR] Excel 동기화 완료")

            except Exception as e:
                logger.error(f"[CACHED_MGR] Excel 동기화 실패: {e}")
                raise

    def _sync_full(self):
        """전체 캐시를 Excel에 덮어쓰기"""
        logger.info("[CACHED_MGR] 전체 동기화 시작...")

        wb = openpyxl.load_workbook(self.excel_filepath)
        ws = wb[self.EXCEL_SHEET_NAME]

        # 기존 데이터 영역 클리어 (rows 54-554)
        for row_idx in range(self.EXCEL_DATA_START_ROW, self.EXCEL_DATA_END_ROW + 1):
            for col_idx in range(1, len(self._column_names) + 1):
                ws.cell(row=row_idx, column=col_idx).value = None

        # 캐시 데이터 쓰기
        for df_idx, row_data in self.cache.iterrows():
            excel_row = self.EXCEL_DATA_START_ROW + df_idx

            if excel_row > self.EXCEL_DATA_END_ROW:
                logger.warning(f"[CACHED_MGR] 데이터 초과: {len(self.cache)}행 > 500행")
                break

            for col_idx, col_name in enumerate(self._column_names, start=1):
                value = row_data.get(col_name)
                ws.cell(row=excel_row, column=col_idx).value = value

        wb.save(self.excel_filepath)
        wb.close()

        logger.info(f"[CACHED_MGR] 전체 동기화 완료: {len(self.cache)}행")

    def _sync_incremental(self):
        """변경사항만 Excel에 반영 (증분 동기화)"""
        pending = self.tracker.get_pending_changes()

        if not pending:
            logger.info("[CACHED_MGR] 동기화할 변경사항 없음")
            return

        logger.info(f"[CACHED_MGR] 증분 동기화 시작: {len(pending)}건")

        wb = openpyxl.load_workbook(self.excel_filepath)
        ws = wb[self.EXCEL_SHEET_NAME]

        trade_id_col = self._find_column('거래ID', 'trade_id')

        for change in pending:
            if change.change_type == ChangeType.CREATE:
                # 빈 행 찾아서 데이터 쓰기
                self._write_to_excel(ws, change.data)

            elif change.change_type == ChangeType.UPDATE:
                # trade_id로 행 찾아서 수정
                self._update_excel_row(ws, change.trade_id, change.data, trade_id_col)

            elif change.change_type == ChangeType.DELETE:
                # trade_id로 행 찾아서 클리어
                self._delete_excel_row(ws, change.trade_id, trade_id_col)

            self.tracker.mark_synced(change)

        wb.save(self.excel_filepath)
        wb.close()

        logger.info(f"[CACHED_MGR] 증분 동기화 완료")

    def _write_to_excel(self, ws, data: Dict):
        """Excel에 새 행 쓰기 (빈 행 찾기)"""
        trade_id_col_idx = self._get_column_index('거래ID', 'trade_id')

        # 빈 행 찾기
        for row_idx in range(self.EXCEL_DATA_START_ROW, self.EXCEL_DATA_END_ROW + 1):
            if ws.cell(row=row_idx, column=trade_id_col_idx).value is None:
                # 데이터 쓰기
                for col_idx, col_name in enumerate(self._column_names, start=1):
                    value = data.get(col_name)
                    ws.cell(row=row_idx, column=col_idx).value = value
                return

        logger.warning("[CACHED_MGR] 빈 행을 찾을 수 없습니다 (500행 초과)")

    def _update_excel_row(self, ws, trade_id: str, data: Dict, trade_id_col: str):
        """Excel 행 업데이트"""
        trade_id_col_idx = self._get_column_index('거래ID', 'trade_id')

        # trade_id로 행 찾기
        for row_idx in range(self.EXCEL_DATA_START_ROW, self.EXCEL_DATA_END_ROW + 1):
            if ws.cell(row=row_idx, column=trade_id_col_idx).value == trade_id:
                # 데이터 업데이트
                for col_name, value in data.items():
                    col_idx = self._get_column_index_by_name(col_name)
                    if col_idx:
                        ws.cell(row=row_idx, column=col_idx).value = value
                return

    def _delete_excel_row(self, ws, trade_id: str, trade_id_col: str):
        """Excel 행 삭제 (None으로 클리어)"""
        trade_id_col_idx = self._get_column_index('거래ID', 'trade_id')

        # trade_id로 행 찾기
        for row_idx in range(self.EXCEL_DATA_START_ROW, self.EXCEL_DATA_END_ROW + 1):
            if ws.cell(row=row_idx, column=trade_id_col_idx).value == trade_id:
                # 행 전체 클리어
                for col_idx in range(1, len(self._column_names) + 1):
                    ws.cell(row=row_idx, column=col_idx).value = None
                return

    def sync_from_excel(self):
        """
        Excel → 캐시 동기화 (사용자가 Excel에서 직접 수정한 경우)

        타임스탬프 기반 충돌 해결:
        - Excel 수정일시 > 캐시 수정일시 → Excel 우선
        - 그 외 → 캐시 유지
        """
        with self.lock:
            try:
                logger.info("[CACHED_MGR] Excel→캐시 동기화 시작...")

                # Excel에서 데이터 로드
                excel_df = self._load_excel_data()

                trade_id_col = self._find_column('거래ID', 'trade_id')
                updated_at_col = self._find_column('수정일시', 'updated_at')

                if not trade_id_col:
                    logger.error("[CACHED_MGR] 거래ID 컬럼을 찾을 수 없습니다")
                    return

                # Excel → 캐시 반영
                for _, excel_row in excel_df.iterrows():
                    excel_trade_id = excel_row.get(trade_id_col)
                    if not excel_trade_id:
                        continue

                    cache_mask = self.cache[trade_id_col] == excel_trade_id

                    if not cache_mask.any():
                        # 신규 행 (Excel에서 추가됨)
                        new_row = pd.DataFrame([excel_row])
                        self.cache = pd.concat([self.cache, new_row], ignore_index=True)
                        logger.info(f"[CACHED_MGR] Excel에서 신규 추가: {excel_trade_id}")

                    else:
                        # 기존 행 (타임스탬프 비교)
                        cache_row = self.cache[cache_mask].iloc[0]

                        if updated_at_col:
                            cache_time = pd.to_datetime(cache_row.get(updated_at_col), errors='coerce')
                            excel_time = pd.to_datetime(excel_row.get(updated_at_col), errors='coerce')

                            if pd.notna(excel_time) and (pd.isna(cache_time) or excel_time > cache_time):
                                # Excel이 더 최신 → 캐시 업데이트
                                idx = cache_mask.idxmax()
                                self.cache.loc[idx] = excel_row
                                logger.info(f"[CACHED_MGR] Excel 우선 업데이트: {excel_trade_id}")

                # Excel에서 삭제된 행 감지
                excel_ids = set(excel_df[trade_id_col].dropna())
                cache_ids = set(self.cache[trade_id_col].dropna())
                deleted_ids = cache_ids - excel_ids

                for trade_id in deleted_ids:
                    self.cache = self.cache[self.cache[trade_id_col] != trade_id].reset_index(drop=True)
                    logger.info(f"[CACHED_MGR] Excel에서 삭제 감지: {trade_id}")

                logger.info("[CACHED_MGR] Excel→캐시 동기화 완료")

            except Exception as e:
                logger.error(f"[CACHED_MGR] Excel→캐시 동기화 실패: {e}")
                raise

    def _load_excel_data(self) -> pd.DataFrame:
        """Excel 데이터 영역만 로드"""
        wb = openpyxl.load_workbook(self.excel_filepath, data_only=True)
        ws = wb[self.EXCEL_SHEET_NAME]

        data_rows = []
        for row_idx in range(self.EXCEL_DATA_START_ROW, self.EXCEL_DATA_END_ROW + 1):
            row_data = {}
            is_empty = True

            for col_idx, col_name in enumerate(self._column_names, start=1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None and value != '':
                    is_empty = False
                row_data[col_name] = value

            if not is_empty:
                data_rows.append(row_data)

        wb.close()

        return pd.DataFrame(data_rows, columns=self._column_names)

    def _find_column(self, *candidates: str) -> Optional[str]:
        """
        컬럼명 찾기 (여러 후보 중 첫번째 매치)

        Args:
            *candidates: 찾을 컬럼명 후보들

        Returns:
            찾은 컬럼명 또는 None
        """
        for col in self._column_names:
            # 컬럼명 정규화 (첫 줄만)
            normalized = self.mapper.normalize_column_name(col)

            for candidate in candidates:
                if candidate in col or candidate in normalized:
                    return col

        return None

    def _get_column_index(self, *candidates: str) -> Optional[int]:
        """컬럼 인덱스 찾기 (1-based)"""
        col_name = self._find_column(*candidates)
        if col_name and col_name in self._column_names:
            return self._column_names.index(col_name) + 1
        return None

    def _get_column_index_by_name(self, col_name: str) -> Optional[int]:
        """정확한 컬럼명으로 인덱스 찾기 (1-based)"""
        if col_name in self._column_names:
            return self._column_names.index(col_name) + 1
        return None

    def get_statistics(self) -> Dict[str, Any]:
        """통계 정보 반환"""
        with self.lock:
            tracker_stats = self.tracker.get_statistics()

            return {
                'cache_rows': len(self.cache) if self.cache is not None else 0,
                'excel_capacity': self.EXCEL_DATA_END_ROW - self.EXCEL_DATA_START_ROW + 1,
                'pending_changes': tracker_stats['pending'],
                'total_changes': tracker_stats['total'],
                'creates': tracker_stats['creates'],
                'updates': tracker_stats['updates'],
                'deletes': tracker_stats['deletes'],
            }

    def __repr__(self):
        stats = self.get_statistics()
        return (f"CachedMasterDataManager("
                f"rows={stats['cache_rows']}, "
                f"pending={stats['pending_changes']}, "
                f"total_changes={stats['total_changes']})")
