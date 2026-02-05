# -*- coding: utf-8 -*-
"""
trade_erp_master_template.xlsx 연동 매니저
- PAGE1_DATA: 거래 데이터 입력/관리 (39개 컬럼)
- PAGE2_VIEW: 집계 조회 (SUMIFS 기반)

요구사항:
1. 수입/수출 관리의 '등록' 버튼 → PAGE1_DATA에 자동 입력
2. 대시보드 월별 상세 실적 → PAGE2 데이터 + 엑셀 다운로드
3. 거래 목록 → PAGE1_DATA와 실시간 연동
"""
import logging
import os
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List, Optional, Tuple
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import shutil

logger = logging.getLogger(__name__)


class TemplateExcelManager:
    """
    trade_erp_master_template.xlsx 전용 관리자
    
    구조:
    - PAGE1_DATA: Row 53 = 헤더, Row 54-554 = 데이터 (500행)
    - PAGE2_VIEW: 집계 필터 + SUMIFS 결과
    """
    
    SHEET_DATA = "PAGE1_DATA"
    SHEET_VIEW = "PAGE2_VIEW"
    HEADER_ROW = 53
    DATA_START_ROW = 54
    DATA_END_ROW = 554
    MAX_ROWS = 500
    
    # PAGE1_DATA 컬럼 매핑 (영문키 → 엑셀 헤더)
    COLUMN_MAPPING = {
        'trade_id': '거래ID\n(trade_id)',
        'direction': '수입/수출\n(direction)',
        'trade_date': '거래일\n(trade_date)',
        'status': '상태\n(status)',
        'item_line_no': '라인\n(item_line_no)',
        'item_name': '물품명\n(item_name)',
        'hscode': 'HS\n(hscode)',
        'origin_country': '원산지\n(origin_country)',
        'importer_name': '수입회사\n(importer_name)',
        'exporter_name': '수출회사\n(exporter_name)',
        'import_country': '수입국\n(import_country)',
        'export_country': '수출국\n(export_country)',
        'incoterms': '인코텀즈\n(incoterms)',
        'currency': '통화\n(currency)',
        'unit_price': '단가\n(unit_price)',
        'quantity': '수량\n(quantity)',
        'uom': '단위\n(uom)',
        'line_amount': '라인금액\n(line_amount)',
        'freight': '운임\n(freight)',
        'insurance': '보험\n(insurance)',
        'invoice_total': '인보이스총액\n(invoice_total)',
        'ci_no': 'C/I\n(ci_no)',
        'pl_no': 'P/L\n(pl_no)',
        'bl_no': 'B/L\n(bl_no)',
        'loading_port': 'POL\n(loading_port)',
        'discharge_port': 'POD\n(discharge_port)',
        'vessel': '선박/편명\n(vessel)',
        'shipment_date': '선적일\n(shipment_date)',
        'discharge_date': '하역일\n(discharge_date)',
        'etd': 'ETD\n(etd)',
        'eta': 'ETA\n(eta)',
        'gross_weight': 'G.W.\n(gross_weight)',
        'net_weight': 'N.W.\n(net_weight)',
        'customs_decl_no': '신고번호\n(customs_decl_no)',
        'fta_applicable': 'FTA\n(fta_applicable)',
        'source_module': '출처\n(source_module)',
        'source_doc_type': '서류유형\n(source_doc_type)',
        'created_at': '생성일시\n(created_at)',
        'updated_at': '수정일시\n(updated_at)',
        # 추가된 컬럼 (관리용)
        'is_important': '중요\n(is_important)',
        'notes': '메모\n(notes)',
    }
    
    # 역매핑
    REVERSE_MAPPING = {v: k for k, v in COLUMN_MAPPING.items()}
    
    def __init__(self, template_path: str = None):
        """
        Args:
            template_path: trade_erp_master_template.xlsx 경로
        """
        if template_path is None:
            base_dir = Path(__file__).parent.parent.parent
            template_path = base_dir / "trade_erp_master_template.xlsx"
        
        self.template_path = Path(template_path)
        
        if not self.template_path.exists():
            raise FileNotFoundError(f"템플릿 파일이 없습니다: {self.template_path}")
        
        # 컬럼 인덱스 캐시
        self._column_indices: Dict[str, int] = {}
        self._load_column_indices()
        
        logger.info(f"[TEMPLATE_MGR] 초기화 완료: {self.template_path}")
    
    def _load_column_indices(self):
        """헤더 행에서 컬럼 인덱스 로드"""
        wb = load_workbook(self.template_path, data_only=True)
        ws = wb[self.SHEET_DATA]
        
        for col_idx in range(1, 50):
            cell_value = ws.cell(row=self.HEADER_ROW, column=col_idx).value
            if cell_value:
                # 영문키로 변환
                eng_key = self.REVERSE_MAPPING.get(cell_value)
                if eng_key:
                    self._column_indices[eng_key] = col_idx
                # 원본 헤더로도 저장
                self._column_indices[cell_value] = col_idx
            else:
                break
        
        wb.close()
        logger.info(f"[TEMPLATE_MGR] 컬럼 인덱스 로드: {len(self._column_indices)}개")
    
    def _get_col_idx(self, key: str) -> Optional[int]:
        """컬럼 키로 인덱스 가져오기"""
        if key in self._column_indices:
            return self._column_indices[key]
        # 영문 → 한글 매핑 시도
        korean_header = self.COLUMN_MAPPING.get(key)
        if korean_header and korean_header in self._column_indices:
            return self._column_indices[korean_header]
        return None
    
    # =========================================================
    # PAGE1_DATA CRUD 메서드
    # =========================================================
    
    def read_all_trades(self) -> pd.DataFrame:
        """
        PAGE1_DATA에서 모든 거래 데이터 읽기
        
        Returns:
            DataFrame (영문 컬럼명 사용)
        """
        wb = load_workbook(self.template_path, data_only=True)
        ws = wb[self.SHEET_DATA]
        
        # 헤더 읽기
        headers = []
        eng_headers = []
        for col_idx in range(1, 50):
            cell_value = ws.cell(row=self.HEADER_ROW, column=col_idx).value
            if cell_value:
                headers.append(cell_value)
                eng_key = self.REVERSE_MAPPING.get(cell_value, cell_value)
                eng_headers.append(eng_key)
            else:
                break
        
        # 데이터 읽기
        data_rows = []
        for row_idx in range(self.DATA_START_ROW, self.DATA_END_ROW + 1):
            row_data = {}
            is_empty = True
            
            for col_idx, eng_key in enumerate(eng_headers, start=1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None and str(value).strip() != '':
                    is_empty = False
                row_data[eng_key] = value
            
            if not is_empty:
                data_rows.append(row_data)
        
        wb.close()
        
        if data_rows:
            df = pd.DataFrame(data_rows, columns=eng_headers)
        else:
            df = pd.DataFrame(columns=eng_headers)
        
        logger.info(f"[TEMPLATE_MGR] 데이터 로드: {len(df)}건")
        return df
    
    def create_trade(self, trade_type: str, data: Dict[str, Any]) -> str:
        """
        새 거래 생성 (PAGE1_DATA에 행 추가)
        
        Args:
            trade_type: 'import' 또는 'export'
            data: 거래 데이터 (영문 키)
        
        Returns:
            생성된 trade_id
        """
        wb = load_workbook(self.template_path)
        ws = wb[self.SHEET_DATA]
        
        # trade_id 생성
        prefix = "IMP" if trade_type == "import" else "EXP"
        date_str = datetime.now().strftime('%Y%m%d')
        
        # 기존 ID 확인
        trade_id_col = self._get_col_idx('trade_id')
        existing_ids = []
        for row_idx in range(self.DATA_START_ROW, self.DATA_END_ROW + 1):
            val = ws.cell(row=row_idx, column=trade_id_col).value
            if val and f"{prefix}-{date_str}" in str(val):
                existing_ids.append(val)
        
        seq = len(existing_ids) + 1
        trade_id = f"{prefix}-{date_str}-{seq:03d}"
        
        # 빈 행 찾기
        target_row = None
        for row_idx in range(self.DATA_START_ROW, self.DATA_END_ROW + 1):
            if ws.cell(row=row_idx, column=trade_id_col).value is None:
                target_row = row_idx
                break
        
        if target_row is None:
            wb.close()
            raise ValueError("데이터 행이 가득 찼습니다 (최대 500행)")
        
        # 데이터 준비
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        final_data = {
            'trade_id': trade_id,
            'direction': '수입' if trade_type == 'import' else '수출',
            'trade_date': data.get('trade_date', datetime.now().strftime('%Y-%m-%d')),
            'status': data.get('status', 'open'),
            'item_line_no': data.get('item_line_no', 1),
            'item_name': data.get('item_name', ''),
            'hscode': data.get('hs_code', data.get('hscode', '')),
            'origin_country': data.get('origin_country', ''),
            'importer_name': data.get('importer_name', data.get('import_company', '')),
            'exporter_name': data.get('exporter_name', data.get('export_company', '')),
            'import_country': data.get('import_country', ''),
            'export_country': data.get('export_country', ''),
            'incoterms': data.get('incoterms', 'FOB'),
            'currency': data.get('currency', 'USD'),
            'unit_price': data.get('unit_price', 0),
            'quantity': data.get('quantity', 0),
            'uom': data.get('uom', data.get('unit', 'EA')),
            'line_amount': data.get('line_amount', data.get('item_value', 0)),
            'freight': data.get('freight', 0),
            'insurance': data.get('insurance', 0),
            'invoice_total': data.get('invoice_total', data.get('total_amount', 0)),
            'ci_no': data.get('ci_no', ''),
            'pl_no': data.get('pl_no', ''),
            'bl_no': data.get('bl_no', data.get('bl_number', '')),
            'loading_port': data.get('loading_port', data.get('port_of_loading', '')),
            'discharge_port': data.get('discharge_port', data.get('port_of_discharge', '')),
            'vessel': data.get('vessel', data.get('vessel_name', '')),
            'shipment_date': data.get('shipment_date', ''),
            'discharge_date': data.get('discharge_date', ''),
            'etd': data.get('etd', ''),
            'eta': data.get('eta', ''),
            'gross_weight': data.get('gross_weight', 0),
            'net_weight': data.get('net_weight', 0),
            'customs_decl_no': data.get('customs_decl_no', ''),
            'fta_applicable': data.get('fta_applicable', 'N'),
            'source_module': data.get('source_module', '수입관리' if trade_type == 'import' else '수출관리'),
            'source_doc_type': data.get('source_doc_type', ''),
            'created_at': now,
            'updated_at': now,
        }
        
        # line_amount 자동 계산
        if not final_data['line_amount'] or final_data['line_amount'] == 0:
            unit_price = float(final_data['unit_price'] or 0)
            quantity = float(final_data['quantity'] or 0)
            final_data['line_amount'] = unit_price * quantity
        
        # 데이터 쓰기
        for key, value in final_data.items():
            col_idx = self._get_col_idx(key)
            if col_idx:
                ws.cell(row=target_row, column=col_idx).value = value
        
        wb.save(self.template_path)
        wb.close()
        
        logger.info(f"[TEMPLATE_MGR] 거래 생성: {trade_id}")
        return trade_id
    
    def update_trade(self, trade_id: str, data: Dict[str, Any]) -> bool:
        """
        거래 수정
        
        Args:
            trade_id: 거래 ID
            data: 수정할 데이터
        
        Returns:
            성공 여부
        """
        wb = load_workbook(self.template_path)
        ws = wb[self.SHEET_DATA]
        
        trade_id_col = self._get_col_idx('trade_id')
        target_row = None
        
        for row_idx in range(self.DATA_START_ROW, self.DATA_END_ROW + 1):
            if ws.cell(row=row_idx, column=trade_id_col).value == trade_id:
                target_row = row_idx
                break
        
        if target_row is None:
            wb.close()
            return False
        
        # 수정일시 업데이트
        data['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # 데이터 쓰기
        for key, value in data.items():
            col_idx = self._get_col_idx(key)
            if col_idx:
                ws.cell(row=target_row, column=col_idx).value = value
        
        wb.save(self.template_path)
        wb.close()
        
        logger.info(f"[TEMPLATE_MGR] 거래 수정: {trade_id}")
        return True
    
    def delete_trade(self, trade_id: str) -> bool:
        """
        거래 삭제 (행 클리어)
        
        Args:
            trade_id: 거래 ID
        
        Returns:
            성공 여부
        """
        wb = load_workbook(self.template_path)
        ws = wb[self.SHEET_DATA]
        
        trade_id_col = self._get_col_idx('trade_id')
        target_row = None
        
        for row_idx in range(self.DATA_START_ROW, self.DATA_END_ROW + 1):
            if ws.cell(row=row_idx, column=trade_id_col).value == trade_id:
                target_row = row_idx
                break
        
        if target_row is None:
            wb.close()
            return False
        
        # 행 클리어
        for col_idx in range(1, len(self._column_indices) // 2 + 1):
            ws.cell(row=target_row, column=col_idx).value = None
        
        wb.save(self.template_path)
        wb.close()
        
        logger.info(f"[TEMPLATE_MGR] 거래 삭제: {trade_id}")
        return True
    
    def get_trade(self, trade_id: str) -> Optional[Dict]:
        """
        단일 거래 조회
        
        Args:
            trade_id: 거래 ID
        
        Returns:
            거래 데이터 또는 None
        """
        df = self.read_all_trades()
        matching = df[df['trade_id'] == trade_id]
        
        if matching.empty:
            return None
        
        return matching.iloc[0].to_dict()
    
    # =========================================================
    # PAGE2_VIEW 관련 메서드
    # =========================================================
    
    def get_monthly_summary(
        self,
        year: int = None,
        month: int = None,
        item_name: str = None,
        import_country: str = None,
        export_country: str = None,
        origin_country: str = None
    ) -> pd.DataFrame:
        """
        PAGE2_VIEW 스타일 월별 집계
        
        Args:
            year: 년도 필터
            month: 월 필터
            item_name: 물품명 필터
            import_country: 수입국 필터
            export_country: 수출국 필터
            origin_country: 원산지 필터
        
        Returns:
            월별 수입액/수출액/순액 DataFrame
        """
        df = self.read_all_trades()
        
        if df.empty:
            return pd.DataFrame(columns=['month', 'import', 'export', 'net_sales'])
        
        # 거래일을 datetime으로 변환
        df['trade_date'] = pd.to_datetime(df['trade_date'], errors='coerce')
        df = df.dropna(subset=['trade_date'])
        
        if df.empty:
            return pd.DataFrame(columns=['month', 'import', 'export', 'net_sales'])
        
        # 년/월 컬럼 생성
        df['year'] = df['trade_date'].dt.year
        df['month_num'] = df['trade_date'].dt.month
        
        # 필터 적용
        if year:
            df = df[df['year'] == year]
        if month:
            df = df[df['month_num'] == month]
        if item_name:
            df = df[df['item_name'].str.contains(item_name, na=False, case=False)]
        if import_country:
            df = df[df['import_country'] == import_country]
        if export_country:
            df = df[df['export_country'] == export_country]
        if origin_country:
            df = df[df['origin_country'] == origin_country]
        
        if df.empty:
            return pd.DataFrame(columns=['month', 'import', 'export', 'net_sales'])
        
        # line_amount 숫자로 변환
        df['line_amount'] = pd.to_numeric(df['line_amount'], errors='coerce').fillna(0)
        
        # 수입/수출 분리
        import_df = df[df['direction'] == '수입'].groupby(
            df['trade_date'].dt.to_period('M')
        )['line_amount'].sum()
        
        export_df = df[df['direction'] == '수출'].groupby(
            df['trade_date'].dt.to_period('M')
        )['line_amount'].sum()
        
        # 결과 DataFrame 생성
        all_months = set(import_df.index) | set(export_df.index)
        
        result = []
        for month_period in sorted(all_months):
            imp_val = import_df.get(month_period, 0)
            exp_val = export_df.get(month_period, 0)
            result.append({
                'month': month_period.to_timestamp(),
                'import': imp_val,
                'export': exp_val,
                'net_sales': exp_val - imp_val
            })
        
        return pd.DataFrame(result)
    
    def get_filter_options(self) -> Dict[str, List[str]]:
        """
        필터 옵션 목록 가져오기 (드롭다운용)
        
        Returns:
            {필드명: 고유값 리스트}
        """
        df = self.read_all_trades()
        
        if df.empty:
            return {
                'item_names': [],
                'import_countries': [],
                'export_countries': [],
                'origin_countries': [],
                'years': [],
                'months': list(range(1, 13))
            }
        
        # 거래일에서 년도 추출
        df['trade_date'] = pd.to_datetime(df['trade_date'], errors='coerce')
        years = sorted(df['trade_date'].dt.year.dropna().unique().astype(int).tolist())
        
        return {
            'item_names': sorted(df['item_name'].dropna().unique().tolist()),
            'import_countries': sorted(df['import_country'].dropna().unique().tolist()),
            'export_countries': sorted(df['export_country'].dropna().unique().tolist()),
            'origin_countries': sorted(df['origin_country'].dropna().unique().tolist()),
            'years': years if years else [datetime.now().year],
            'months': list(range(1, 13))
        }
    
    # =========================================================
    # 엑셀 다운로드 메서드
    # =========================================================
    
    def get_template_for_download(self) -> str:
        """
        현재 trade_erp_master_template.xlsx 파일 경로 반환 (다운로드용)
        
        Returns:
            파일 경로
        """
        return str(self.template_path)
    
    def export_monthly_to_excel(self, monthly_df: pd.DataFrame, output_path: str = None) -> str:
        """
        월별 실적 데이터를 별도 Excel로 내보내기
        
        Args:
            monthly_df: 월별 데이터 DataFrame
            output_path: 출력 경로 (없으면 자동 생성)
        
        Returns:
            생성된 파일 경로
        """
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_dir = self.template_path.parent / "data" / "exports"
            output_dir.mkdir(parents=True, exist_ok=True)
            output_path = output_dir / f"monthly_summary_{timestamp}.xlsx"
        
        monthly_df.to_excel(output_path, index=False, engine='openpyxl')
        logger.info(f"[TEMPLATE_MGR] 월별 실적 내보내기: {output_path}")
        return str(output_path)
    
    # =========================================================
    # 통계 메서드
    # =========================================================
    
    def get_statistics(self) -> Dict[str, Any]:
        """
        전체 통계 정보
        
        Returns:
            통계 딕셔너리
        """
        df = self.read_all_trades()
        
        if df.empty:
            return {
                'total': 0,
                'import_count': 0,
                'export_count': 0,
                'total_import_amount': 0,
                'total_export_amount': 0,
                'unique_items': 0,
                'unique_countries': 0,
            }
        
        df['line_amount'] = pd.to_numeric(df['line_amount'], errors='coerce').fillna(0)
        
        import_df = df[df['direction'] == '수입']
        export_df = df[df['direction'] == '수출']
        
        return {
            'total': len(df),
            'import_count': len(import_df),
            'export_count': len(export_df),
            'total_import_amount': import_df['line_amount'].sum(),
            'total_export_amount': export_df['line_amount'].sum(),
            'unique_items': df['item_name'].nunique(),
            'unique_countries': len(set(df['import_country'].dropna()) | set(df['export_country'].dropna())),
        }


# 싱글톤 인스턴스 (편의용)
_template_manager: Optional[TemplateExcelManager] = None


def get_template_manager(template_path: str = None) -> TemplateExcelManager:
    """템플릿 매니저 싱글톤 인스턴스 가져오기"""
    global _template_manager
    
    if _template_manager is None:
        _template_manager = TemplateExcelManager(template_path)
    
    return _template_manager


def reset_template_manager():
    """템플릿 매니저 리셋 (재로드 필요 시)"""
    global _template_manager
    _template_manager = None
