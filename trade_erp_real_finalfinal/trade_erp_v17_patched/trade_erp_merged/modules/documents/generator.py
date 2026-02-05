# -*- coding: utf-8 -*-
"""
서류 생성 모듈 v2.0
- Commercial Invoice (상공회의소 양식)
- Packing List (상공회의소 양식)
- 수입신고서 (관세청 양식)
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, Optional, List
import shutil

import sys
sys.path.insert(0, str(Path(__file__).parent.parent.parent))
from config.settings import TEMPLATES_DIR, PROCESSED_DATA_DIR

logger = logging.getLogger(__name__)


class DocumentGenerator:
    """서류 생성기 v2.0"""
    
    def __init__(self):
        self.output_dir = PROCESSED_DATA_DIR / "documents"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.templates_dir = TEMPLATES_DIR
    
    def generate_commercial_invoice(self, trade_data: Dict[str, Any]) -> str:
        """Commercial Invoice 생성 (상공회의소 양식)"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("openpyxl 미설치")
            return ""
        
        template_path = self.templates_dir / "COMMERCIAL_INVOICE_template.xlsx"
        if not template_path.exists():
            logger.error(f"템플릿 파일 없음: {template_path}")
            return ""
        
        # 템플릿 복사
        filename = f"CI_{trade_data.get('trade_id', 'draft')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        output_path = self.output_dir / filename
        shutil.copy(template_path, output_path)
        
        # 데이터 입력
        wb = load_workbook(output_path)
        ws = wb.active
        
        # 셀 매핑 (상공회의소 양식 기준)
        # 1. SHIPPER/EXPORTER (B4 아래)
        ws['B5'] = trade_data.get('exporter_name', '')
        ws['B6'] = trade_data.get('exporter_address', '')
        ws['B7'] = trade_data.get('exporter_tel', '')
        
        # 8. NO. & DATE OF INVOICE (H6 아래)
        ws['H7'] = f"{trade_data.get('trade_id', '')} / {datetime.now().strftime('%Y-%m-%d')}"
        
        # 2. FOR ACCOUNT AND RISK OF MESSRS (B9 아래 - Buyer/Importer)
        ws['B10'] = trade_data.get('importer_name', '')
        ws['B11'] = trade_data.get('importer_address', '')
        ws['B12'] = trade_data.get('importer_tel', '')
        
        # 9. NO. & DATE OF L/C (H9 아래)
        ws['H10'] = trade_data.get('lc_number', '')
        
        # 10. L/C ISSUING BANK (H12 아래)
        ws['H13'] = trade_data.get('lc_bank', '')
        
        # 3. NOTIFY PARTY (B14 아래)
        ws['B15'] = trade_data.get('notify_party', trade_data.get('importer_name', ''))
        
        # 4. PORT OF LOADING (B19 아래)
        ws['B20'] = trade_data.get('port_of_loading', '')
        
        # 5. FINAL DESTINATION (E19 아래)
        ws['E20'] = trade_data.get('port_of_discharge', trade_data.get('destination', ''))
        
        # 6. VESSEL/FLIGHT (B22 아래)
        ws['B23'] = trade_data.get('vessel_name', '')
        
        # 7. SAILING ON OR ABOUT (E22 아래)
        ws['E23'] = trade_data.get('sailing_date', '')
        
        # 11. REMARKS (H18 아래)
        ws['H19'] = trade_data.get('remarks', '')
        
        # 물품 정보 (Row 26부터 시작)
        row = 26
        currency = trade_data.get('currency', 'USD')
        
        # 12. MARKS AND NUMBERS (B열)
        ws[f'B{row}'] = trade_data.get('marks', 'N/M')
        
        # 13. DESCRIPTIONS OF GOODS (E열)
        ws[f'E{row}'] = trade_data.get('item_name', '')
        ws[f'E{row+1}'] = f"HS Code: {trade_data.get('hs_code', '')}"
        
        # 14. QUANTITY/UNIT (H열)
        ws[f'H{row}'] = f"{trade_data.get('quantity', '')} {trade_data.get('unit', 'PCS')}"
        
        # 15. UNIT-PRICE (J열)
        ws[f'J{row}'] = f"{currency} {trade_data.get('unit_price', 0):,.2f}"
        
        # 16. AMOUNT (L열)
        total_amount = trade_data.get('item_value', trade_data.get('unit_price', 0) * trade_data.get('quantity', 1))
        ws[f'L{row}'] = f"{currency} {total_amount:,.2f}"
        
        # TOTAL (Row 38)
        ws['L38'] = f"{currency} {total_amount:,.2f}"
        
        # 17. SIGNED BY (H43)
        ws['H43'] = trade_data.get('signatory', '')
        
        wb.save(output_path)
        logger.info(f"[DOC] Commercial Invoice 생성: {output_path}")
        return str(output_path)
    
    def generate_packing_list(self, trade_data: Dict[str, Any]) -> str:
        """Packing List 생성 (상공회의소 양식)"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("openpyxl 미설치")
            return ""
        
        template_path = self.templates_dir / "PACKING_LIST_template.xlsx"
        if not template_path.exists():
            logger.error(f"템플릿 파일 없음: {template_path}")
            return ""
        
        # 템플릿 복사
        filename = f"PL_{trade_data.get('trade_id', 'draft')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        output_path = self.output_dir / filename
        shutil.copy(template_path, output_path)
        
        # 데이터 입력
        wb = load_workbook(output_path)
        ws = wb.active
        
        # 1. SHIPPER/EXPORTER (B4 아래)
        ws['B5'] = trade_data.get('exporter_name', '')
        ws['B6'] = trade_data.get('exporter_address', '')
        ws['B7'] = trade_data.get('exporter_tel', '')
        
        # 8. NO. & DATE OF INVOICE (H6 아래)
        ws['H7'] = f"PL-{trade_data.get('trade_id', '')} / {datetime.now().strftime('%Y-%m-%d')}"
        
        # 2. TO APPLICANT/CONSIGNEE (B9 아래)
        ws['B10'] = trade_data.get('importer_name', '')
        ws['B11'] = trade_data.get('importer_address', '')
        ws['B12'] = trade_data.get('importer_tel', '')
        
        # 3. NOTIFY PARTY (B14 아래)
        ws['B15'] = trade_data.get('notify_party', trade_data.get('importer_name', ''))
        
        # 4. PORT OF LOADING (B19 아래)
        ws['B20'] = trade_data.get('port_of_loading', '')
        
        # 5. FINAL DESTINATION (E19 아래)
        ws['E20'] = trade_data.get('port_of_discharge', trade_data.get('destination', ''))
        
        # 6. VESSEL NAME (B22 아래)
        ws['B23'] = trade_data.get('vessel_name', '')
        
        # 7. SAILING ON OR ABOUT (E22 아래)
        ws['E23'] = trade_data.get('sailing_date', '')
        
        # 9. REMARKS (H9 아래)
        ws['H10'] = trade_data.get('remarks', '')
        
        # 물품 정보 (Row 26부터 시작)
        row = 26
        
        # 10. MARKS AND NUMBERS (B열)
        ws[f'B{row}'] = trade_data.get('marks', 'N/M')
        
        # 11. DESCRIPTIONS OF GOODS (E열)
        ws[f'E{row}'] = trade_data.get('item_name', '')
        ws[f'E{row+1}'] = f"HS Code: {trade_data.get('hs_code', '')}"
        
        # 12. QUANTITY (H열)
        ws[f'H{row}'] = f"{trade_data.get('quantity', '')} {trade_data.get('unit', 'PCS')}"
        
        # 13. NET-WEIGHT (J열)
        ws[f'J{row}'] = f"{trade_data.get('net_weight', '')} KG"
        
        # 14. GROSS-WEIGHT (L열)
        ws[f'L{row}'] = f"{trade_data.get('gross_weight', '')} KG"
        
        # TOTAL (Row 44)
        ws['H44'] = f"{trade_data.get('quantity', '')} {trade_data.get('unit', 'PCS')}"
        ws['J44'] = f"{trade_data.get('net_weight', '')} KG"
        ws['L44'] = f"{trade_data.get('gross_weight', '')} KG"
        
        wb.save(output_path)
        logger.info(f"[DOC] Packing List 생성: {output_path}")
        return str(output_path)
    
    def generate_import_declaration(self, trade_data: Dict[str, Any]) -> str:
        """수입신고서 생성 (관세청 양식)"""
        try:
            from docx import Document
        except ImportError:
            logger.error("python-docx 미설치")
            return ""
        
        template_path = self.templates_dir / "수입신고서_template.docx"
        if not template_path.exists():
            logger.error(f"템플릿 파일 없음: {template_path}")
            return ""
        
        # 템플릿 복사
        filename = f"수입신고서_{trade_data.get('trade_id', 'draft')}_{datetime.now().strftime('%Y%m%d')}.docx"
        output_path = self.output_dir / filename
        shutil.copy(template_path, output_path)
        
        # 데이터 입력
        doc = Document(output_path)
        
        # 테이블 내 텍스트 치환
        replacements = {
            '99999-99-9999999-9': trade_data.get('declaration_no', ''),
            'YYYY/MM/DD': datetime.now().strftime('%Y/%m/%d'),
            'XXXXXXXXXXXXXXXX(X': trade_data.get('bl_number', ''),
            'YYXXXXXXXXXX-9999-999': trade_data.get('cargo_management_no', ''),
            '⑩신 고 자 XXXXXXX': f"⑩신 고 자 {trade_data.get('declarant', '')}",
            '⑪수 입 자 XXXXXXXX': f"⑪수 입 자 {trade_data.get('importer_name', '')}",
            '품 명 XXXXXXXXXXXXXXXXXXXXXXXXXX': f"품 명 {trade_data.get('item_name', '')}",
            '상 표 XXXXXXXXXXXXXXXXXXXXXXXXXX': f"상 표 {trade_data.get('brand', '')}",
            '9999.99-9999': trade_data.get('hs_code', ''),
            '$999,999,999,999': f"${trade_data.get('cif_value_foreign', 0):,.0f}",
            '\\999,999,999,999': f"₩{trade_data.get('cif_value_krw', 0):,.0f}",
            'XX XXXXXXXXXXXX': f"{trade_data.get('origin_country_code', '')} {trade_data.get('origin_country', '')}",
            'XXXXXXXXXXXXXXXX XX': trade_data.get('vessel_name', ''),
        }
        
        for table in doc.tables:
            for row in table.rows:
                for cell in row.cells:
                    for old_text, new_text in replacements.items():
                        if old_text in cell.text:
                            cell.text = cell.text.replace(old_text, str(new_text))
        
        doc.save(output_path)
        logger.info(f"[DOC] 수입신고서 생성: {output_path}")
        return str(output_path)


def generate_all_documents(trade_data: Dict[str, Any], trade_type: str = 'import') -> List[Dict[str, Any]]:
    """전체 서류 일괄 생성"""
    gen = DocumentGenerator()
    generated = []
    
    # Commercial Invoice
    path = gen.generate_commercial_invoice(trade_data)
    if path:
        generated.append({'name': 'Commercial Invoice', 'path': path})
    
    # Packing List
    path = gen.generate_packing_list(trade_data)
    if path:
        generated.append({'name': 'Packing List', 'path': path})
    
    # 수입신고서 (수입인 경우만)
    if trade_type == 'import':
        path = gen.generate_import_declaration(trade_data)
        if path:
            generated.append({'name': '수입신고서', 'path': path})
    
    return generated